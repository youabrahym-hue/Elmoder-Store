import os, sqlite3, hashlib, shutil, secrets, json, io
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, abort

APP_NAME='Elmoder-Store'; SUBTITLE='Kings Of Apples'
BASE_DIR=os.path.dirname(os.path.abspath(__file__))
DATA_DIR=os.path.join(BASE_DIR,'data'); BACKUP_DIR=os.path.join(DATA_DIR,'backups')
os.makedirs(BACKUP_DIR, exist_ok=True)
DB_FILE=os.path.join(DATA_DIR,'elmoder_store.db')
DATABASE_URL=os.environ.get('DATABASE_URL','').strip()
USE_POSTGRES=bool(DATABASE_URL)
app=Flask(__name__)
app.secret_key=os.environ.get('ELMODER_SECRET_KEY') or secrets.token_hex(32)

I18N={
'en': {'dashboard':'Dashboard','devices':'Devices','reports':'Reports','users':'Users','backup':'Backup & Restore','settings':'Settings','logout':'Logout','login':'Login','username':'Username','password':'Password','remember':'Remember me','updated':'All device statuses have been updated.','total_devices':'Total Devices','available':'Available','sold':'Sold','sales':'Sales','capital':'Capital','profit':'Profit','loss':'Loss','net_profit':'Net Profit','add_device':'Add Device','edit':'Edit','delete':'Delete','save':'Save','cancel':'Cancel','clear':'Clear','merchant':'Merchant','device_name':'Device Name','model':'Model','imei':'IMEI','tax_status':'Tax Status','purchase_aed':'Purchase AED','purchase_egp':'Purchase EGP','uae_expenses':'UAE Expenses','traveler':'Traveler','receiving':'Receiving Egypt','sale_price':'Sale Price','status':'Status','invoice':'Invoice','search':'Search device, IMEI, model, merchant...','add_merchant':'Add Merchant','add_device_name':'Add Device Name','add_user':'Add User','role':'Role','active':'Active','employee':'Employee','admin':'Admin','permissions':'Permissions','create_backup':'Create Backup','download_backup':'Download Backup','restore_backup':'Restore Backup','appearance':'Appearance','language':'Language','light':'Light','dark':'Dark','blue':'Blue','green':'Green','orange':'Orange','red':'Red','save_settings':'Save Settings','no_permission':'You do not have permission.','invalid_login':'Invalid username or password.','required':'Device name is required.','saved':'Saved successfully.','deleted':'Deleted successfully.','invoice_details':'Invoice Details','date':'Date','actions':'Actions','activity':'Recent Activity','top_device':'Top Device','top_merchant':'Top Merchant'},
'ar': {'dashboard':'لوحة التحكم','devices':'الأجهزة','reports':'التقارير','users':'المستخدمون','backup':'النسخ الاحتياطي','settings':'الإعدادات','logout':'تسجيل الخروج','login':'تسجيل الدخول','username':'اسم المستخدم','password':'كلمة المرور','remember':'تذكرني','updated':'تم تحديث حالة جميع الأجهزة','total_devices':'إجمالي الأجهزة','available':'متاحة','sold':'مباعة','sales':'المبيعات','capital':'رأس المال','profit':'الربح','loss':'الخسائر','net_profit':'صافي الربح','add_device':'إضافة جهاز','edit':'تعديل','delete':'حذف','save':'حفظ','cancel':'إلغاء','clear':'مسح','merchant':'التاجر','device_name':'اسم الجهاز','model':'الموديل','imei':'IMEI','tax_status':'الحالة الضريبية','purchase_aed':'سعر الشراء بالدرهم','purchase_egp':'سعر الشراء بالمصري','uae_expenses':'المصاريف بالإمارات','traveler':'المسافر','receiving':'مصاريف الاستلام بمصر','sale_price':'سعر البيع','status':'الحالة','invoice':'الفاتورة','search':'ابحث بالجهاز أو IMEI أو الموديل أو التاجر...','add_merchant':'إضافة تاجر','add_device_name':'إضافة اسم جهاز','add_user':'إضافة مستخدم','role':'الصلاحية','active':'نشط','employee':'موظف','admin':'مدير','permissions':'الصلاحيات','create_backup':'إنشاء نسخة احتياطية','download_backup':'تحميل النسخة','restore_backup':'استعادة نسخة','appearance':'المظهر','language':'اللغة','light':'فاتح','dark':'داكن','blue':'أزرق','green':'أخضر','orange':'برتقالي','red':'أحمر','save_settings':'حفظ الإعدادات','no_permission':'ليس لديك صلاحية لتنفيذ هذا الإجراء.','invalid_login':'اسم المستخدم أو كلمة المرور غير صحيحة.','required':'اسم الجهاز مطلوب.','saved':'تم الحفظ بنجاح.','deleted':'تم الحذف بنجاح.','invoice_details':'تفاصيل الفاتورة','date':'التاريخ','actions':'الإجراءات','activity':'آخر الأنشطة','top_device':'أكثر جهاز','top_merchant':'أكثر تاجر'}}

def h(p): return hashlib.sha256(p.encode()).hexdigest()
def now(): return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
class PGConnection:
    def __init__(self, url):
        import psycopg
        from psycopg.rows import dict_row
        self.con=psycopg.connect(url, row_factory=dict_row)
    def _sql(self, sql):
        sql=sql.replace('INSERT OR IGNORE INTO', 'INSERT INTO')
        if 'INSERT INTO' in sql and 'ON CONFLICT' not in sql and 'SELECT' not in sql:
            # Safe for the small lookup tables used by this app.
            if any(t in sql for t in ('merchants','device_names')):
                sql=sql.rstrip(';')+' ON CONFLICT DO NOTHING'
        return sql.replace('?', '%s')
    def execute(self, sql, params=()):
        cur=self.con.cursor(); cur.execute(self._sql(sql), params); return cur
    def executescript(self, script):
        self.con.execute(script)
    def commit(self): self.con.commit()
    def close(self): self.con.close()

def db():
    if USE_POSTGRES:
        return PGConnection(DATABASE_URL)
    con=sqlite3.connect(DB_FILE); con.row_factory=sqlite3.Row; return con

def init_db():
    con=db()
    if USE_POSTGRES:
        con.executescript('''
        CREATE TABLE IF NOT EXISTS users(id BIGSERIAL PRIMARY KEY,username TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'Employee',status TEXT NOT NULL DEFAULT 'Active',can_view_devices INTEGER DEFAULT 1,can_add_devices INTEGER DEFAULT 1,can_edit_devices INTEGER DEFAULT 1,can_delete_devices INTEGER DEFAULT 0,can_view_profit INTEGER DEFAULT 0,can_manage_users INTEGER DEFAULT 0,can_manage_settings INTEGER DEFAULT 0,can_manage_backup INTEGER DEFAULT 0,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS devices(id BIGSERIAL PRIMARY KEY,date TEXT NOT NULL,invoice_no TEXT,merchant TEXT,device_name TEXT NOT NULL,model TEXT,imei TEXT,tax_status TEXT DEFAULT 'معفي',purchase_aed DOUBLE PRECISION DEFAULT 0,purchase_egp DOUBLE PRECISION DEFAULT 0,uae_expenses DOUBLE PRECISION DEFAULT 0,traveler DOUBLE PRECISION DEFAULT 0,receiving DOUBLE PRECISION DEFAULT 0,sale_price DOUBLE PRECISION DEFAULT 0,profit DOUBLE PRECISION DEFAULT 0,loss DOUBLE PRECISION DEFAULT 0,status TEXT DEFAULT 'Available',created_at TEXT NOT NULL,updated_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS merchants(id BIGSERIAL PRIMARY KEY,name TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS device_names(id BIGSERIAL PRIMARY KEY,name TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS activity_log(id BIGSERIAL PRIMARY KEY,username TEXT,action TEXT NOT NULL,details TEXT,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS notifications(id BIGSERIAL PRIMARY KEY,message TEXT NOT NULL,kind TEXT DEFAULT 'info',is_read INTEGER DEFAULT 0,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
        CREATE TABLE IF NOT EXISTS backups(id BIGSERIAL PRIMARY KEY,filename TEXT UNIQUE NOT NULL,payload TEXT NOT NULL,created_at TEXT NOT NULL);
        ''')
    else:
        con.executescript('''
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'Employee',status TEXT NOT NULL DEFAULT 'Active',can_view_devices INTEGER DEFAULT 1,can_add_devices INTEGER DEFAULT 1,can_edit_devices INTEGER DEFAULT 1,can_delete_devices INTEGER DEFAULT 0,can_view_profit INTEGER DEFAULT 0,can_manage_users INTEGER DEFAULT 0,can_manage_settings INTEGER DEFAULT 0,can_manage_backup INTEGER DEFAULT 0,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS devices(id INTEGER PRIMARY KEY AUTOINCREMENT,date TEXT NOT NULL,invoice_no TEXT,merchant TEXT,device_name TEXT NOT NULL,model TEXT,imei TEXT,tax_status TEXT DEFAULT 'معفي',purchase_aed REAL DEFAULT 0,purchase_egp REAL DEFAULT 0,uae_expenses REAL DEFAULT 0,traveler REAL DEFAULT 0,receiving REAL DEFAULT 0,sale_price REAL DEFAULT 0,profit REAL DEFAULT 0,loss REAL DEFAULT 0,status TEXT DEFAULT 'Available',created_at TEXT NOT NULL,updated_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS merchants(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS device_names(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS activity_log(id INTEGER PRIMARY KEY AUTOINCREMENT,username TEXT,action TEXT NOT NULL,details TEXT,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS notifications(id INTEGER PRIMARY KEY AUTOINCREMENT,message TEXT NOT NULL,kind TEXT DEFAULT 'info',is_read INTEGER DEFAULT 0,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
        CREATE TABLE IF NOT EXISTS backups(id INTEGER PRIMARY KEY AUTOINCREMENT,filename TEXT UNIQUE NOT NULL,payload TEXT NOT NULL,created_at TEXT NOT NULL);
        ''')
    con.commit()
    if not con.execute('SELECT id FROM users LIMIT 1').fetchone():
        con.execute('INSERT INTO users(username,password_hash,role,status,can_view_devices,can_add_devices,can_edit_devices,can_delete_devices,can_view_profit,can_manage_users,can_manage_settings,can_manage_backup,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',('admin',h('1234'),'Admin','Active',1,1,1,1,1,1,1,1,now())); con.commit()
    con.close()

def setting(k, default=None):
    con=db(); r=con.execute('SELECT value FROM settings WHERE key=?',(k,)).fetchone(); con.close(); return r['value'] if r else default
def set_setting(k,v):
    con=db(); con.execute('INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value',(k,str(v))); con.commit(); con.close()

def current_user():
    uid=session.get('uid')
    if not uid:return None
    con=db(); r=con.execute('SELECT * FROM users WHERE id=?',(uid,)).fetchone(); con.close(); return r

def login_required(f):
    @wraps(f)
    def w(*a,**kw):
        if not current_user(): return redirect(url_for('login'))
        return f(*a,**kw)
    return w

def perm(name):
    u=current_user(); return bool(u and (u['role']=='Admin' or u[name]))

@app.context_processor
def ctx():
    lang=session.get('lang') or setting('language','ar') or 'ar'; theme=session.get('theme') or setting('theme','light') or 'light'; color=session.get('color') or setting('color','blue') or 'blue'
    return dict(t=lambda k:I18N.get(lang,I18N['ar']).get(k,k),lang=lang,theme=theme,color=color,user=current_user())


@app.get('/health')
def health():
    return {'status':'ok','database':'postgres' if USE_POSTGRES else 'sqlite'}

@app.route('/',methods=['GET'])
def index(): return redirect(url_for('dashboard') if current_user() else url_for('login'))
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        con=db(); r=con.execute("SELECT * FROM users WHERE username=? AND password_hash=? AND status='Active'",(request.form.get('username','').strip(),h(request.form.get('password','')))).fetchone(); con.close()
        if r:
            session['uid']=r['id']; session['lang']=setting('language','ar'); session['theme']=setting('theme','light'); session['color']=setting('color','blue'); flash(I18N[session['lang']]['updated'],'success'); return redirect(url_for('dashboard'))
        flash(I18N['ar']['invalid_login'],'danger')
    return render_template('login.html')
@app.route('/logout')
def logout(): session.clear(); return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    con=db(); s=con.execute("SELECT COUNT(*) total,SUM(CASE WHEN status='Available' THEN 1 ELSE 0 END) available,SUM(CASE WHEN status='Sold' THEN 1 ELSE 0 END) sold,COALESCE(SUM(sale_price),0) sales,COALESCE(SUM(purchase_egp+uae_expenses+traveler+receiving),0) capital,COALESCE(SUM(profit),0) profit,COALESCE(SUM(loss),0) loss FROM devices").fetchone(); logs=con.execute('SELECT * FROM activity_log ORDER BY id DESC LIMIT 8').fetchall(); topd=con.execute("SELECT device_name,COUNT(*) c FROM devices WHERE sale_price>0 GROUP BY device_name ORDER BY c DESC LIMIT 1").fetchone(); topm=con.execute("SELECT merchant,COUNT(*) c FROM devices WHERE merchant<>'' GROUP BY merchant ORDER BY c DESC LIMIT 1").fetchone(); con.close()
    return render_template('dashboard.html',s=s,logs=logs,topd=topd,topm=topm)

@app.route('/devices',methods=['GET','POST'])
@login_required
def devices():
    con=db(); q=request.args.get('q','').strip()
    if request.method=='POST':
        if not perm('can_add_devices'): abort(403)
        f=request.form; name=f.get('device_name','').strip()
        if not name: flash(I18N[session.get('lang','ar')]['required'],'danger'); return redirect(url_for('devices'))
        nums=lambda k: float(f.get(k) or 0)
        purchase,uae,trav,recv,sale,loss=map(nums,['purchase_egp','uae_expenses','traveler','receiving','sale_price','loss']); profit=max(0,sale-(purchase+uae+trav+recv)); status=f.get('status') or ('Sold' if sale>0 else 'Available')
        invoice=f"INV-{datetime.now().year}-"+datetime.now().strftime('%m%d%H%M%S')
        con.execute('INSERT INTO devices(date,invoice_no,merchant,device_name,model,imei,tax_status,purchase_aed,purchase_egp,uae_expenses,traveler,receiving,sale_price,profit,loss,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(datetime.now().strftime('%Y-%m-%d'),invoice,f.get('merchant'),name,f.get('model'),f.get('imei'),f.get('tax_status','معفي'),nums('purchase_aed'),purchase,uae,trav,recv,sale,profit,loss,status,now(),now()))
        con.execute('INSERT OR IGNORE INTO merchants(name) VALUES(?)',(f.get('merchant',''),)) if f.get('merchant') else None
        con.execute('INSERT OR IGNORE INTO device_names(name) VALUES(?)',(name,)); con.execute('INSERT INTO activity_log(username,action,details,created_at) VALUES(?,?,?,?)',(current_user()['username'],'Add Device',name,now())); con.commit(); flash(I18N[session.get('lang','ar')]['saved'],'success'); return redirect(url_for('devices'))
    rows=con.execute('SELECT * FROM devices WHERE merchant LIKE ? OR device_name LIKE ? OR model LIKE ? OR imei LIKE ? ORDER BY id DESC',tuple([f'%{q}%']*4)).fetchall(); merchants=con.execute('SELECT name FROM merchants WHERE name<>"" ORDER BY name').fetchall(); names=con.execute('SELECT name FROM device_names ORDER BY name').fetchall(); con.close()
    return render_template('devices.html',rows=rows,merchants=merchants,names=names)

@app.route('/devices/<int:did>/edit',methods=['GET'])
@login_required
def edit_device_page(did):
    if not perm('can_edit_devices'): abort(403)
    con=db(); r=con.execute('SELECT * FROM devices WHERE id=?',(did,)).fetchone(); merchants=con.execute('SELECT name FROM merchants WHERE name<>"" ORDER BY name').fetchall(); names=con.execute('SELECT name FROM device_names ORDER BY name').fetchall(); con.close()
    if not r: abort(404)
    return render_template('edit_device.html',r=r,merchants=merchants,names=names)

@app.route('/devices/<int:did>/edit',methods=['POST'])
@login_required
def edit_device(did):
    if not perm('can_edit_devices'): abort(403)
    con=db(); r=con.execute('SELECT * FROM devices WHERE id=?',(did,)).fetchone();
    if not r: abort(404)
    f=request.form; nums=lambda k: float(f.get(k) or 0); purchase,uae,trav,recv,sale,loss=map(nums,['purchase_egp','uae_expenses','traveler','receiving','sale_price','loss']); profit=max(0,sale-(purchase+uae+trav+recv)); status=f.get('status') or ('Sold' if sale>0 else 'Available')
    con.execute('UPDATE devices SET merchant=?,device_name=?,model=?,imei=?,tax_status=?,purchase_aed=?,purchase_egp=?,uae_expenses=?,traveler=?,receiving=?,sale_price=?,profit=?,loss=?,status=?,updated_at=? WHERE id=?',(f.get('merchant'),f.get('device_name'),f.get('model'),f.get('imei'),f.get('tax_status'),nums('purchase_aed'),purchase,uae,trav,recv,sale,profit,loss,status,now(),did)); con.execute('INSERT INTO activity_log(username,action,details,created_at) VALUES(?,?,?,?)',(current_user()['username'],'Edit Device',f"#{did}",now())); con.commit(); con.close(); flash(I18N[session.get('lang','ar')]['saved'],'success'); return redirect(url_for('devices'))
@app.route('/devices/<int:did>/delete',methods=['POST'])
@login_required
def delete_device(did):
    if not perm('can_delete_devices'): abort(403)
    con=db(); con.execute('DELETE FROM devices WHERE id=?',(did,)); con.execute('INSERT INTO activity_log(username,action,details,created_at) VALUES(?,?,?,?)',(current_user()['username'],'Delete Device',f"#{did}",now())); con.commit(); con.close(); flash(I18N[session.get('lang','ar')]['deleted'],'success'); return redirect(url_for('devices'))

@app.route('/reports')
@login_required
def reports():
    con=db(); s=con.execute('SELECT COUNT(*) count,COALESCE(SUM(purchase_egp+uae_expenses+traveler+receiving),0) capital,COALESCE(SUM(sale_price),0) sales,COALESCE(SUM(profit),0) profit,COALESCE(SUM(loss),0) loss FROM devices').fetchone(); by=con.execute('SELECT merchant,COUNT(*) count,COALESCE(SUM(sale_price),0) sales,COALESCE(SUM(profit),0) profit FROM devices GROUP BY merchant ORDER BY sales DESC').fetchall(); con.close(); return render_template('reports.html',s=s,by=by)

@app.route('/invoice/<int:did>')
@login_required
def invoice(did):
    con=db(); r=con.execute('SELECT * FROM devices WHERE id=?',(did,)).fetchone(); con.close();
    if not r: abort(404)
    return render_template('invoice.html',r=r)

@app.route('/users',methods=['GET','POST'])
@login_required
def users():
    if not perm('can_manage_users'): abort(403)
    con=db()
    if request.method=='POST':
        f=request.form; username=f.get('username','').strip(); password=f.get('password',''); role=f.get('role','Employee'); status=f.get('status','Active');
        try:
            con.execute('INSERT INTO users(username,password_hash,role,status,can_view_devices,can_add_devices,can_edit_devices,can_delete_devices,can_view_profit,can_manage_users,can_manage_settings,can_manage_backup,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',(username,h(password),role,status,*[1 if f.get(k) else 0 for k in ['can_view_devices','can_add_devices','can_edit_devices','can_delete_devices','can_view_profit','can_manage_users','can_manage_settings','can_manage_backup']],now())); con.commit(); flash(I18N[session.get('lang','ar')]['saved'],'success')
        except Exception as e:
            if 'unique' in str(e).lower(): flash('Username already exists.','danger')
            else: raise
    rows=con.execute('SELECT * FROM users ORDER BY id').fetchall(); con.close(); return render_template('users.html',rows=rows)
@app.route('/users/<int:uid>/edit',methods=['GET','POST'])
@login_required
def edit_user(uid):
    if not perm('can_manage_users'): abort(403)
    con=db(); r=con.execute('SELECT * FROM users WHERE id=?',(uid,)).fetchone()
    if not r: abort(404)
    if request.method=='POST':
        f=request.form; username=f.get('username','').strip(); role=f.get('role','Employee'); status=f.get('status','Active'); perms=[1 if f.get(k) else 0 for k in ['can_view_devices','can_add_devices','can_edit_devices','can_delete_devices','can_view_profit','can_manage_users','can_manage_settings','can_manage_backup']]
        if f.get('password'): con.execute('UPDATE users SET username=?,password_hash=?,role=?,status=?,can_view_devices=?,can_add_devices=?,can_edit_devices=?,can_delete_devices=?,can_view_profit=?,can_manage_users=?,can_manage_settings=?,can_manage_backup=? WHERE id=?',(username,h(f.get('password')),role,status,*perms,uid))
        else: con.execute('UPDATE users SET username=?,role=?,status=?,can_view_devices=?,can_add_devices=?,can_edit_devices=?,can_delete_devices=?,can_view_profit=?,can_manage_users=?,can_manage_settings=?,can_manage_backup=? WHERE id=?',(username,role,status,*perms,uid))
        con.commit(); con.close(); flash(I18N[session.get('lang','ar')]['saved'],'success'); return redirect(url_for('users'))
    con.close(); return render_template('edit_user.html',r=r)

@app.route('/users/<int:uid>/delete',methods=['POST'])
@login_required
def delete_user(uid):
    if not perm('can_manage_users'): abort(403)
    if uid==current_user()['id']: abort(400)
    con=db(); con.execute('DELETE FROM users WHERE id=?',(uid,)); con.commit(); con.close(); return redirect(url_for('users'))

def excel_book(title, headers, rows, totals=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb=Workbook(); ws=wb.active; ws.title=title[:31]
    ws.append(headers)
    header_fill=PatternFill('solid', fgColor='2563EB')
    for c in ws[1]:
        c.font=Font(bold=True,color='FFFFFF')
        c.fill=header_fill
        c.alignment=Alignment(horizontal='center',vertical='center')
    for row in rows: ws.append(list(row))
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; ws.sheet_view.showGridLines=False
    for col in ws.columns:
        letter=get_column_letter(col[0].column)
        width=min(max(max(len(str(x.value or '')) for x in col)+2,10),32)
        ws.column_dimensions[letter].width=width
    if totals:
        r=ws.max_row+2
        for label, formula in totals:
            ws.cell(r,1,label); ws.cell(r,2,formula); r+=1
    return wb

def send_excel(wb, filename):
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _excel_value(row, *keys):
    for k in keys:
        if k in row and row[k] is not None:
            v=str(row[k]).strip()
            if v != '': return v
    return ''

def _num_value(row, *keys):
    v=_excel_value(row,*keys)
    if not v: return 0.0
    try: return float(str(v).replace(',',''))
    except (TypeError,ValueError): return 0.0

def _norm_header(value):
    return str(value or '').strip().lower().replace(' ','_').replace('-','_')

def _import_excel_file(uploaded):
    from openpyxl import load_workbook
    wb=load_workbook(uploaded, read_only=True, data_only=True)
    ws=wb.active
    values=list(ws.iter_rows(values_only=True))
    if not values: return 0,0
    raw_headers=[_norm_header(x) for x in values[0]]
    aliases={
        'date':['date','التاريخ'], 'invoice_no':['invoice','invoice_no','رقم_الفاتورة'],
        'merchant':['merchant','التاجر'], 'device_name':['device','device_name','اسم_الجهاز'],
        'model':['model','الموديل'], 'imei':['imei'], 'tax_status':['tax_status','الحالة_الضريبية'],
        'purchase_aed':['purchase_aed','purchase_aed_','سعر_الشراء_بالدرهم'], 'purchase_egp':['purchase_egp','سعر_الشراء_بالمصري'],
        'uae_expenses':['uae_expenses','المصاريف_بالإمارات'], 'traveler':['traveler','المسافر'],
        'receiving':['receiving','receiving_egypt','مصاريف_الاستلام_بمصر'], 'sale_price':['sale_price','سعر_البيع'],
        'profit':['profit','الربح'], 'loss':['loss','الخسائر'], 'status':['status','الحالة']
    }
    index={}
    for dest,names in aliases.items():
        for name in names:
            nn=_norm_header(name)
            if nn in raw_headers: index[dest]=raw_headers.index(nn); break
    if 'device_name' not in index: raise ValueError('ملف Excel يجب أن يحتوي على عمود Device أو اسم الجهاز.')
    con=db(); added=0; skipped=0
    try:
        for vals in values[1:]:
            row={k: (vals[i] if i < len(vals) else '') for k,i in index.items()}
            name=str(row.get('device_name') or '').strip()
            if not name: continue
            imei=str(row.get('imei') or '').strip()
            if imei:
                exists=con.execute('SELECT id FROM devices WHERE imei=? LIMIT 1',(imei,)).fetchone()
                if exists: skipped+=1; continue
            purchase=_num_value(row,'purchase_egp'); uae=_num_value(row,'uae_expenses'); trav=_num_value(row,'traveler'); recv=_num_value(row,'receiving'); sale=_num_value(row,'sale_price')
            supplied_profit=_num_value(row,'profit'); loss=_num_value(row,'loss')
            profit=supplied_profit if supplied_profit else max(0,sale-(purchase+uae+trav+recv))
            status=str(row.get('status') or '').strip() or ('Sold' if sale>0 else 'Available')
            date_val=row.get('date'); date_str=str(date_val)[:10] if date_val else datetime.now().strftime('%Y-%m-%d')
            invoice=str(row.get('invoice_no') or '').strip() or f"INV-{datetime.now().year}-"+datetime.now().strftime('%m%d%H%M%S%f')[:12]
            merchant=str(row.get('merchant') or '').strip()
            con.execute('INSERT INTO devices(date,invoice_no,merchant,device_name,model,imei,tax_status,purchase_aed,purchase_egp,uae_expenses,traveler,receiving,sale_price,profit,loss,status,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(date_str,invoice,merchant,name,str(row.get('model') or '').strip(),imei,str(row.get('tax_status') or 'معفي'),_num_value(row,'purchase_aed'),purchase,uae,trav,recv,sale,profit,loss,status,now(),now()))
            if merchant: con.execute('INSERT OR IGNORE INTO merchants(name) VALUES(?)',(merchant,))
            con.execute('INSERT OR IGNORE INTO device_names(name) VALUES(?)',(name,))
            added+=1
        con.execute('INSERT INTO activity_log(username,action,details,created_at) VALUES(?,?,?,?)',(current_user()['username'],'Import Excel',f'Added {added}; skipped {skipped}',now()))
        con.commit()
    finally:
        con.close()
    return added,skipped

@app.route('/import/excel',methods=['POST'])
@login_required
def import_excel():
    if not perm('can_add_devices'): abort(403)
    uploaded=request.files.get('excel_file')
    if not uploaded or not uploaded.filename.lower().endswith('.xlsx'):
        flash('اختر ملف Excel بصيغة .xlsx','danger'); return redirect(url_for('devices'))
    try:
        added,skipped=_import_excel_file(uploaded)
        flash(f'تم استيراد {added} جهاز' + (f' وتخطي {skipped} مكرر' if skipped else ''),'success')
    except Exception as e:
        flash(f'فشل استيراد Excel: {e}','danger')
    return redirect(url_for('devices'))

@app.route('/export/excel/open')
@login_required
def open_excel():
    # Browsers cannot directly launch desktop Excel; provide the same live workbook download.
    return export_devices_excel()

@app.route('/export/excel/devices')
@login_required
def export_devices_excel():
    con=db(); rows=con.execute('SELECT id,date,invoice_no,merchant,device_name,model,imei,tax_status,purchase_aed,purchase_egp,uae_expenses,traveler,receiving,sale_price,profit,loss,status FROM devices ORDER BY id DESC').fetchall(); con.close()
    headers=['ID','Date','Invoice','Merchant','Device','Model','IMEI','Tax Status','Purchase AED','Purchase EGP','UAE Expenses','Traveler','Receiving Egypt','Sale Price','Profit','Loss','Status']
    wb=excel_book('Devices',headers,rows,[('Total Profit',f'=SUM(O2:O{len(rows)+1})'),('Total Loss',f'=SUM(P2:P{len(rows)+1})'),('Net Profit',f'=B{len(rows)+3}-B{len(rows)+4}')])
    return send_excel(wb,f'Elmoder-Store_Devices_{datetime.now():%Y-%m-%d}.xlsx')

@app.route('/export/excel/sales')
@login_required
def export_sales_excel():
    con=db(); rows=con.execute("SELECT id,date,invoice_no,merchant,device_name,model,imei,purchase_egp,uae_expenses,traveler,receiving,sale_price,profit,loss FROM devices WHERE sale_price>0 ORDER BY date DESC,id DESC").fetchall(); con.close()
    headers=['ID','Date','Invoice','Merchant','Device','Model','IMEI','Purchase EGP','UAE Expenses','Traveler','Receiving Egypt','Sale Price','Profit','Loss']
    wb=excel_book('Sales',headers,rows,[('Total Sales',f'=SUM(L2:L{len(rows)+1})'),('Total Profit',f'=SUM(M2:M{len(rows)+1})'),('Total Loss',f'=SUM(N2:N{len(rows)+1})')])
    return send_excel(wb,f'Elmoder-Store_Sales_{datetime.now():%Y-%m-%d}.xlsx')

@app.route('/export/excel/purchases')
@login_required
def export_purchases_excel():
    con=db(); rows=con.execute('SELECT id,date,merchant,device_name,model,imei,purchase_aed,purchase_egp,uae_expenses,traveler,receiving,status FROM devices ORDER BY date DESC,id DESC').fetchall(); con.close()
    headers=['ID','Date','Merchant','Device','Model','IMEI','Purchase AED','Purchase EGP','UAE Expenses','Traveler','Receiving Egypt','Status']
    wb=excel_book('Purchases',headers,rows,[('Total Purchase EGP',f'=SUM(H2:H{len(rows)+1})'),('Total UAE Expenses',f'=SUM(I2:I{len(rows)+1})'),('Total Traveler',f'=SUM(J2:J{len(rows)+1})'),('Total Receiving',f'=SUM(K2:K{len(rows)+1})')])
    return send_excel(wb,f'Elmoder-Store_Purchases_{datetime.now():%Y-%m-%d}.xlsx')

@app.route('/export/excel/full')
@login_required
def export_full_excel():
    if not perm('can_manage_backup'): abort(403)
    con=db()
    devices=con.execute('SELECT * FROM devices ORDER BY id').fetchall()
    merchants=con.execute('SELECT * FROM merchants ORDER BY id').fetchall()
    users=con.execute('SELECT id,username,role,status,created_at FROM users ORDER BY id').fetchall()
    activity=con.execute('SELECT * FROM activity_log ORDER BY id').fetchall()
    settings=con.execute('SELECT * FROM settings ORDER BY key').fetchall()
    con.close()
    wb=excel_book('Devices',[k for k in devices[0].keys()] if devices else ['No data'],devices)
    def add_sheet(name, headers, rows):
        ws=wb.create_sheet(name); ws.append(headers)
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='2563EB'); c.alignment=Alignment(horizontal='center')
        for row in rows: ws.append(list(row))
        ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions; ws.sheet_view.showGridLines=False
        for col in ws.columns: ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(max(len(str(x.value or '')) for x in col)+2,10),32)
    add_sheet('Merchants',['ID','Name'],merchants)
    add_sheet('Users',['ID','Username','Role','Status','Created At'],users)
    add_sheet('Activity',['ID','Username','Action','Details','Created At'],activity)
    add_sheet('Settings',['Key','Value'],settings)
    return send_excel(wb,f'Elmoder-Store_Full_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx')

@app.route('/backup')
@login_required
def backup():
    if not perm('can_manage_backup'): abort(403)
    con=db(); rows=con.execute('SELECT id,filename,created_at FROM backups ORDER BY id DESC').fetchall(); con.close()
    return render_template('backup.html',files=rows,cloud=USE_POSTGRES)

def backup_payload():
    tables=['users','devices','merchants','device_names','activity_log','notifications','settings']
    con=db(); data={}
    for table in tables:
        rows=con.execute(f'SELECT * FROM {table}').fetchall()
        data[table]=[dict(r) for r in rows]
    con.close(); return {'app':'Elmoder-Store','version':2,'created_at':now(),'tables':data}

def _restore_payload(payload):
    if not isinstance(payload,dict) or payload.get('app')!='Elmoder-Store' or not isinstance(payload.get('tables'),dict):
        raise ValueError('ملف النسخة الاحتياطية غير صالح.')
    allowed=['users','devices','merchants','device_names','activity_log','notifications','settings']
    con=db()
    try:
        for table in allowed:
            rows=payload['tables'].get(table,[])
            if not isinstance(rows,list): continue
            con.execute(f'DELETE FROM {table}')
            for row in rows:
                if not isinstance(row,dict) or not row: continue
                keys=list(row.keys()); placeholders=','.join('?' for _ in keys)
                vals=[row[k] for k in keys]
                con.execute(f"INSERT INTO {table} ({','.join(keys)}) VALUES ({placeholders})",vals)
        con.commit()
    finally:
        con.close()

@app.route('/backup/create',methods=['POST'])
@login_required
def backup_create():
    if not perm('can_manage_backup'): abort(403)
    stamp=datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    name=f'elmoder_store_{stamp}.json'; payload=backup_payload(); text=json.dumps(payload,ensure_ascii=False,indent=2,default=str)
    con=db(); con.execute('INSERT INTO backups(filename,payload,created_at) VALUES(?,?,?)',(name,text,now()));
    # Keep the most recent 20 backups in the online database.
    old=con.execute('SELECT id FROM backups ORDER BY id DESC').fetchall()
    for r in old[20:]: con.execute('DELETE FROM backups WHERE id=?',(r['id'],))
    con.commit(); con.close()
    flash('تم إنشاء النسخة الاحتياطية بنجاح','success'); return redirect(url_for('backup'))

@app.route('/backup/download/<int:bid>')
@login_required
def backup_download(bid):
    if not perm('can_manage_backup'): abort(403)
    con=db(); r=con.execute('SELECT filename,payload FROM backups WHERE id=?',(bid,)).fetchone(); con.close()
    if not r: abort(404)
    return send_file(io.BytesIO(r['payload'].encode('utf-8')),as_attachment=True,download_name=r['filename'],mimetype='application/json')

@app.route('/backup/restore',methods=['POST'])
@login_required
def backup_restore():
    if not perm('can_manage_backup'): abort(403)
    uploaded=request.files.get('backup_file')
    if not uploaded or not uploaded.filename.lower().endswith('.json'):
        flash('اختر ملف نسخة احتياطية بصيغة .json','danger'); return redirect(url_for('backup'))
    try:
        payload=json.loads(uploaded.read().decode('utf-8'))
        _restore_payload(payload)
        flash('تمت استعادة النسخة الاحتياطية بنجاح','success')
    except Exception as e:
        flash(f'فشل الاستعادة: {e}','danger')
    return redirect(url_for('backup'))

@app.route('/backup/delete/<int:bid>',methods=['POST'])
@login_required
def backup_delete(bid):
    if not perm('can_manage_backup'): abort(403)
    con=db(); con.execute('DELETE FROM backups WHERE id=?',(bid,)); con.commit(); con.close()
    flash('تم حذف النسخة الاحتياطية','success'); return redirect(url_for('backup'))

@app.route('/settings',methods=['GET','POST'])
@login_required
def settings():
    if not perm('can_manage_settings'): abort(403)
    if request.method=='POST':
        for k in ('language','theme','color'): set_setting(k,request.form.get(k))
        session['lang']=request.form.get('language','ar'); session['theme']=request.form.get('theme','light'); session['color']=request.form.get('color','blue'); flash(I18N[session['lang']]['saved'],'success'); return redirect(url_for('settings'))
    return render_template('settings.html')

@app.route('/lookup/<kind>',methods=['POST'])
@login_required
def lookup(kind):
    if not perm('can_add_devices'): abort(403)
    name=request.form.get('name','').strip();
    if name:
        con=db(); table='merchants' if kind=='merchant' else 'device_names'; con.execute(f'INSERT OR IGNORE INTO {table}(name) VALUES(?)',(name,)); con.commit(); con.close()
    return redirect(url_for('devices'))

@app.errorhandler(403)
def forbidden(e): return ('Forbidden',403)

if __name__=='__main__':
    init_db(); app.run(host='0.0.0.0',port=int(os.environ.get('PORT',5000)),debug=False)
else:
    init_db()
